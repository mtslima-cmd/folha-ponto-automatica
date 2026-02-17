import calendar
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def traduzir_dia(dia_em_ingles):
    traducoes = {
        "Monday": "Segunda-feira",
        "Tuesday": "Terça-feira",
        "Wednesday": "Quarta-feira",
        "Thursday": "Quinta-feira",
        "Friday": "Sexta-feira",
        "Saturday": "Sábado",
        "Sunday": "Domingo",
    }
    return traducoes.get(dia_em_ingles, dia_em_ingles)


def gerar_calendario(mes: int, ano: int):
    calendario = calendar.monthcalendar(ano, mes)

    dias_do_mes = []

    for semana in calendario:
        for dia in semana:
            if dia != 0:
                data = datetime(ano, mes, dia)
                dia_semana_en = data.strftime("%A")
                dia_semana_pt = traduzir_dia(dia_semana_en)

                eh_final_de_semana = dia_semana_pt in ["Sábado", "Domingo"]

                dias_do_mes.append({
                    "data": data,
                    "dia_semana": dia_semana_pt,
                    "final_de_semana": eh_final_de_semana
                })

    return dias_do_mes

def gerar_excel(nome_arquivo, dados):
    wb = Workbook()
    ws = wb.active
    ws.title = "Folha de Ponto"

    cabecalho = [
        "Data",
        "Dia da Semana",
        "Entrada Manhã",
        "Saída Manhã",
        "Entrada Tarde",
        "Saída Tarde",
        "Total de Horas",
        "Observação"
    ]
    ws.append(cabecalho)

    # Estilos
    header_fill = PatternFill("solid", fgColor="1f2937")  # escuro
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    weekend_fill = PatternFill("solid", fgColor="E5E7EB")  # cinza claro
    holiday_fill = PatternFill("solid", fgColor="FEF3C7")  # amarelo claro

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Cabeçalho
    for col in range(1, len(cabecalho) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    # Dados
    for linha in dados:
        ws.append(linha)

    # Ajustes visuais
    ws.freeze_panes = "A2"  # trava cabeçalho

    col_widths = {
        1: 12,  # Data
        2: 16,  # Dia semana
        3: 14,  # Entrada manhã
        4: 12,  # Saída manhã
        5: 14,  # Entrada tarde
        6: 12,  # Saída tarde
        7: 14,  # Total
        8: 22,  # Observação
    }
    for col_idx, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Aplicar bordas, alinhamento e cores por linha
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        obs = (ws.cell(row=r, column=8).value or "").strip()

        # alinhamentos
        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=2).alignment = left

        for col in range(3, 8):  # horários + total
            ws.cell(row=r, column=col).alignment = center

        ws.cell(row=r, column=8).alignment = left

        # bordas em tudo
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = border

        # cor por tipo
        fill = None
        if obs == "Final de Semana":
            fill = weekend_fill
        elif obs in ("Feriado", "Ponto Facultativo", "Meio Expediente"):
            fill = holiday_fill

        if fill:
            for col in range(1, 9):
                ws.cell(row=r, column=col).fill = fill

    # Soma mensal (em horas) – coloca no final
    # Observação: seus totais estão como texto "08:00"/"04:00".
    # Vamos transformar isso em tempo do Excel: =TIME(8,0,0) etc. futuramente.
    # Por enquanto, soma apenas quantidade de dias cheios/parciais:
    # - conta 8h quando Total="08:00"
    # - conta 4h quando Total="04:00"
    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "TOTAL (h)", "", ""])

    total_row = ws.max_row
    ws.cell(row=total_row, column=6).font = Font(bold=True)
    ws.cell(row=total_row, column=6).alignment = right = Alignment(horizontal="right", vertical="center")

    # Fórmula: (conta 08:00)*8 + (conta 04:00)*4
    # Coluna G = 7
    ws.cell(row=total_row, column=7).value = (
        f'=COUNTIF(G2:G{max_row},"08:00")*8 + COUNTIF(G2:G{max_row},"04:00")*4'
    )
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    ws.cell(row=total_row, column=7).alignment = center

    wb.save(nome_arquivo)